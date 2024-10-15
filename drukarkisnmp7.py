import datetime
import openpyxl
import subprocess
from pysnmp.hlapi import *

def ping_device(ip):
    try:
        output = subprocess.check_output(['ping', '-c', '1', ip])
        if '1 packets transmitted, 1 received' in output.decode('utf-8'):
            return True
        else:
            return False
    except subprocess.CalledProcessError:
        return False

def get_printer_info(host, ip):
    if not ping_device(ip):
        return None, f"Ping nie odpowiada"
    
    errorIndication, errorStatus, errorIndex, varBinds = next(
        getCmd(SnmpEngine(),
               CommunityData('PaSwPrD', mpModel=0),
               UdpTransportTarget((ip, 161)),
               ContextData(),
               ObjectType(ObjectIdentity('1.3.6.1.2.1.43.10.2.1.4.1.1')),  # Printer page counter
               ObjectType(ObjectIdentity('1.3.6.1.2.1.43.5.1.1.17.1')),  # Printer serial number
               )
    )

    if errorIndication:
        return None, f"Błąd odczytu: {errorIndication}"
    
    page_counter = int(varBinds[0][1])
    serial_number = str(varBinds[1][1])
    return page_counter, serial_number

def get_last_month_reading(printer_name, worksheet):
    current_date = datetime.date.today()
    last_month = current_date.replace(day=1) - datetime.timedelta(days=1)
    for row in worksheet.iter_rows(values_only=True):
        if row[0] == printer_name and row[2].month == last_month.month and row[2].year == last_month.year:
            return row[3]
    return None

def save_printer_info(printer_info, worksheet, filename):
    current_date = datetime.date.today()
    month_year = current_date.strftime('%B %Y')
    printer_name = printer_info['name']
    previous_reading = get_last_month_reading(printer_name, worksheet)

    if previous_reading is None:
        previous_reading = 0

    if not isinstance(previous_reading, int):
        previous_reading = 0

    monthly_usage = printer_info['page_counter'] - previous_reading

    for row in worksheet.iter_rows(values_only=True):
        if row[0] == printer_name and row[1] == printer_info['ip'] and row[3].month == current_date.month and row[3].year == current_date.year:
            row = list(row)
            if printer_info['page_counter'] is None:
                row[4] = printer_info['error']
                row[5] = None
            else:
                row[4] = printer_info['page_counter']
                row[5] = monthly_usage
            row = tuple(row)
            break
    else:
        if printer_info['page_counter'] is None:
            printer_info['page_counter'] = printer_info['error']
        worksheet.append([printer_name, printer_info['ip'], printer_info['serial_number'], current_date, printer_info['page_counter'], None])



    try:
        worksheet.parent.save(filename)
        return True
    except PermissionError:
        print(f"Nie można zapisać do pliku {filename}. Plik jest zajęty lub nie masz do niego dostępu.")
        answer = input("Czy plik został już zamknięty i jest dostępny do zapisu? (tak/nie): ")
        if answer.lower() == "tak":
            try:
                worksheet.parent.save(filename)
                return True
            except PermissionError:
                new_filename = f"drukarki-wynik_{current_date.strftime('%Y-%m-%d')}.xlsx"
                print(f"Zapisuję dane do nowego pliku: {new_filename}")
                worksheet.parent.save(new_filename)
                return False
        else:
            print("Zakończ program i zamknij plik, aby móc kontynuować zapis.")
            return False


def save_failed_printer_info(printer_info, worksheet):
    current_date = datetime.date.today()
    printer_name = printer_info['name']
    worksheet.append([printer_name, printer_info['ip'], current_date, printer_info['error']])

def read_printers_from_file(file_path):
    printers = []
    with open(file_path, 'r') as file:
        for line in file:
            host, ip = line.strip().split(',')
            printers.append({'name': host, 'ip': ip})
    return printers

def main():
    printers_file = 'drukarki.txt'
    previous_readings_file = 'drukarki-wynik.xlsx'

    printers = read_printers_from_file(printers_file)
    try:
        workbook = openpyxl.load_workbook(previous_readings_file)
        worksheet = workbook.active
    except FileNotFoundError:
        print(f"Plik {previous_readings_file} nie istnieje. Tworzę nowy plik.")
        workbook = openpyxl.Workbook()
        worksheet = workbook.active

    # Checking if there are entries in the sheet
    existing_entries = False
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        existing_entries = True
        break

    answer = ""
    should_override = False  # Variable that specifies whether data should be overwritten for all printers in the same month

    if existing_entries:
        answer = input("Czy nadpisać dane dla wszystkich drukarek w tym samym miesiącu? (tak/nie): ")

    current_month = datetime.date.today().month
    current_year = datetime.date.today().year

    for printer in printers:
        page_counter, serial_number = get_printer_info(printer['name'], printer['ip'])

        if page_counter is None or serial_number is None:
            print(f"Nieudane połączenie z drukarką {printer['name']}. Sprawdź działanie drukarki.")
            printer_info = {
                'name': printer['name'],
                'ip': printer['ip'],
                'error': page_counter if page_counter is None else serial_number
            }
            save_failed_printer_info(printer_info, worksheet)
            continue

        printer_info = {
            'name': printer['name'],
            'ip': printer['ip'],
            'page_counter': page_counter,
            'serial_number': serial_number
        }

        printer_found = False
        for row in worksheet.iter_rows(values_only=True):
            if row[0] == printer_info['name'] and row[1] == printer_info['ip']:
                printer_found = True
                row_month = row[2].month
                row_year = row[2].year
                if row_month == current_month and row_year == current_year and answer.lower() == "tak":
                    row = list(row)
                    row[2] = datetime.date.today()
                    row[3] = printer_info['page_counter']
                    row[4] = None
                    break

        if not printer_found:
            should_override = True  #If a new printer is added, set the should_override flag to True
            worksheet.append([printer_info['name'], printer_info['ip'], datetime.date.today(), printer_info['page_counter'], None])

    if should_override and answer.lower() == "tak":
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            row_month = row[2].month
            row_year = row[2].year

            if row_month == current_month and row_year == current_year:
                printer_name = row[0]
                printer_ip = row[1]
                page_counter, serial_number = get_printer_info(printer_name, printer_ip)

                if page_counter is None or serial_number is None:
                    print(f"Nieudane połączenie z drukarką {printer_name}. Sprawdź działanie drukarki.")
                    printer_info = {
                        'name': printer_name,
                        'ip': printer_ip,
                        'error': page_counter if page_counter is None else serial_number
                    }
                    save_failed_printer_info(printer_info, worksheet)
                    continue

                row = list(row)
                row[3] = page_counter
                row[4] = None

    try:
        workbook.save(previous_readings_file)
        workbook.close()
    except PermissionError:
        print(f"Nie można zapisać do pliku {previous_readings_file}. Plik jest zajęty lub nie masz do niego dostępu.")

if __name__ == '__main__':
    main()
