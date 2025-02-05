# Script for reading data from printers via SNMP - for auditing the page counter ver. 1.0
# Author Kuleszap (Piotr Kulesza)
#
# Futher function : 
# - should be run once a month


import datetime
import csv
import subprocess
import os
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from pysnmp.hlapi import *

# Function to read SNMP data from printer
def read_snmp_data(host, oid):
    errorIndication, errorStatus, errorIndex, varBinds = next(
        getCmd(SnmpEngine(),
               CommunityData('TEST', mpModel=0),
               UdpTransportTarget((host, 161)),
               ContextData(),
               ObjectType(ObjectIdentity(oid)))
    )

    if errorIndication:
        return None
    elif errorStatus:
        return None
    else:
        for varBind in varBinds:
            return varBind.prettyPrint().split('=', 1)[-1].strip().split(' ', 1)[0]

# Function to read data from file
def read_printer_list(file_path):
    printer_list = []
    with open(file_path, 'r') as file:
        reader = csv.reader(file, delimiter=',')
        for row in reader:
            host, ip = row
            printer_list.append((host, ip))
    return printer_list

# Function to check device availability via ping
def check_ping(host):
    result = subprocess.run(['ping', '-c', '1', '-W', '1', host], stdout=subprocess.PIPE, stderr=subprocess.DEVNULL)
    return result.returncode == 0

# Funkcja do zapisu danych do pliku Excel
def save_to_excel(data):
    wb = Workbook()
    ws = wb.active

    ws.append(['Index', 'Host', 'IP', 'Producent', 'Model', 'Serial', 'Licznik', 'Data odczytu (miesiąc-rok)'])

    for index, row in enumerate(data, start=1):
        ws.append([index] + row + [datetime.datetime.now().strftime('%m-%Y')])

    try:
        wb.save('wynik.xlsx')
    except PermissionError:
        current_time = datetime.datetime.now().strftime('%Y%m%d-%H%M%S')
        file_name = f'wynik_{current_time}.xlsx'
        wb.save(file_name)
        print(f"Nie można zapisać pliku wynik.xlsx. Dane zapisano w pliku {file_name}.")

# Main
def main():
    printer_list = read_printer_list('drukarki.txt')
    data = []
    start_time = datetime.datetime.now()

    print("Wybierz tryb działania:")
    print("1. Sprawdzanie pinga, gdy SNMP nie działa.")
    print("2. Bez sprawdzania pinga (szybki tryb).")

    choice = input("Wybierz tryb (1 lub 2): ")

    if choice == '1':
        for host, ip in printer_list:
            manufacturer = read_snmp_data(ip, '1.3.6.1.2.1.1.1.0')
            model = read_snmp_data(ip, '1.3.6.1.2.1.25.3.2.1.3.1')
            serial_number = read_snmp_data(ip, '1.3.6.1.2.1.43.5.1.1.17.1')
            counter = read_snmp_data(ip, '1.3.6.1.2.1.43.10.2.1.4.1.1')

            if manufacturer and model and serial_number and counter:
                data.append([host, ip, manufacturer, model, serial_number, counter])
            else:
                if check_ping(ip):
                    if manufacturer and model and serial_number:
                        data.append([host, ip, manufacturer, model, serial_number, 'Błąd SNMP'])
                    else:
                        data.append([host, ip, 'Błąd drukarki', '', '', ''])
                else:
                    data.append([host, ip, 'Błąd drukarki', '', '', ''])
    elif choice == '2':
        for host, ip in printer_list:
            manufacturer = read_snmp_data(ip, '1.3.6.1.2.1.1.1.0')
            model = read_snmp_data(ip, '1.3.6.1.2.1.25.3.2.1.3.1')
            serial_number = read_snmp_data(ip, '1.3.6.1.2.1.43.5.1.1.17.1')
            counter = read_snmp_data(ip, '1.3.6.1.2.1.43.10.2.1.4.1.1')

            if manufacturer and model and serial_number and counter:
                data.append([host, ip, manufacturer, model, serial_number, counter])
            else:
                data.append([host, ip, 'Błąd drukarki', '', '', ''])

    end_time = datetime.datetime.now()
    execution_time = end_time - start_time
    print(f"Sprawdzono {len(printer_list)} drukarek. Czas wykonania: {execution_time}")

    try:
        save_to_excel(data)
    except Exception as e:
        current_time = datetime.datetime.now().strftime('%Y%m%d-%H%M%S')
        file_name = f'wynik_{current_time}.xlsx'
        print(f"Wystąpił błąd podczas zapisu pliku wynik.xlsx. Dane zapisano w pliku {file_name}.")
        wb = Workbook()
        ws = wb.active
        ws.append(['Błąd podczas zapisu pliku wynik.xlsx'])
        ws.append([str(e)])
        wb.save(file_name)

# Run main function
if __name__ == '__main__':
    main()
